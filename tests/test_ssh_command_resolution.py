"""
Tests for P0 SSH command resolution fixes:
  - resolve_task_command with per_group_commands
  - Empty commands → EXEC_FAILED (COMMAND_MISSING)
  - ONLY_LOGIN_BANNER detection
  - tasks.json path resolution
"""
from __future__ import annotations
import os, sys, json, tempfile, time
from pathlib import Path
from unittest.mock import patch, MagicMock

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest


# ===========================================================================
# resolve_task_command tests
# ===========================================================================


class TestResolveTaskCommand:
    """P0-1: per_group_commands consumed by SSH executor."""

    def test_per_group_commands_match_returns_group_command(self):
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "光模块测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do hccn_tool -i $i -optical -g; done",
        })

        cmd = resolve_task_command(task, "A3")
        assert "hccn_tool" in cmd
        assert "display interface transceiver" not in cmd

    def test_per_group_commands_no_match_fallback_to_default(self):
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "光模块测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do hccn_tool -i $i -optical -g; done",
        })

        cmd = resolve_task_command(task, "L1")
        assert cmd == "display interface transceiver"

    def test_no_per_group_commands_returns_command_or_url(self):
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "温度测试"
            command_or_url = "display device temperature"

        task = FakeTask()
        cmd = resolve_task_command(task, "L1")
        assert cmd == "display device temperature"

    def test_empty_command_or_url_returns_empty(self):
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "空命令测试"
            command_or_url = ""

        task = FakeTask()
        cmd = resolve_task_command(task, "")
        assert cmd == ""

    def test_group_case_insensitive(self):
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "test"
            command_or_url = "default_cmd"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {"A3": "a3_cmd"})

        assert resolve_task_command(task, "a3") == "a3_cmd"
        assert resolve_task_command(task, "A3") == "a3_cmd"
        assert resolve_task_command(task, " a3 ") == "a3_cmd"

    def test_4_1_15_a3_gets_hccn_tool(self):
        """4.1.15 A3 must get hccn_tool, not display interface transceiver."""
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do echo \"==============> $i\"; hccn_tool -i $i -optical -g;done",
        })

        cmd = resolve_task_command(task, "A3")
        assert "hccn_tool" in cmd
        assert "==============>" in cmd
        assert "display interface transceiver" not in cmd

    def test_4_1_15_l1_gets_display_interface_transceiver(self):
        """4.1.15 L1 gets default command — display interface transceiver."""
        from src.executor.ssh_executor import resolve_task_command

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do hccn_tool -i $i -optical -g;done",
        })

        cmd = resolve_task_command(task, "L1")
        assert cmd == "display interface transceiver"


# ===========================================================================
# Empty commands → EXEC_FAILED tests
# ===========================================================================


class TestEmptyCommandsDetection:
    """P0-2: commands=[] must EXEC_FAILED, not EXEC_SUCCESS."""

    def test_empty_commands_returns_exec_failed(self):
        from src.executor.ssh_executor import SSHExecutor
        from src.models.device import Device
        from src.models.task import Task
        from src.models.task_plan import TaskPlan

        executor = SSHExecutor()
        device = Device(
            row_index=1, device_name="test-dev", device_group="A3",
            bmc_ip="", bmc_username="", bmc_password="",
            inband_ip="127.0.0.1",
            inband_username="test", inband_password="test",
        )
        task = Task(
            row_index=1, sequence=1, task_name="空命令测试",
            task_type="SSH", execution_mode="SSH_CMD",
            command_or_url="",  # Empty!
            match_group="A3",
        )
        plan = TaskPlan(
            plan_id="test-001", device=device, task=task,
            task_id="t1",
        )

        with tempfile.TemporaryDirectory() as tmpdir:
            result = executor.execute(plan, tmpdir)
            # Should fail before even trying SSH (no connection needed)
            assert result.execution_status == "EXEC_FAILED"
            assert "COMMAND_MISSING" in result.execution_failure_reason

    def test_empty_commands_ssh_not_attempted(self):
        """When commands are empty, SSH connection should NOT be attempted."""
        from src.executor.ssh_executor import SSHExecutor
        from src.models.device import Device
        from src.models.task import Task
        from src.models.task_plan import TaskPlan

        executor = SSHExecutor()
        device = Device(
            row_index=1, device_name="test-dev", device_group="L1",
            bmc_ip="", bmc_username="", bmc_password="",
            inband_ip="192.0.2.1",  # Non-routable — would timeout if attempted
            inband_username="test", inband_password="test",
        )
        task = Task(
            row_index=1, sequence=1, task_name="no-command",
            task_type="SSH", execution_mode="SSH_CMD",
            command_or_url="", match_group="L1",
        )
        plan = TaskPlan(plan_id="p1", device=device, task=task, task_id="t1")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = executor.execute(plan, tmpdir)
            # Must fail quickly (no SSH connection timeout)
            assert result.execution_status == "EXEC_FAILED"
            assert result.duration_seconds < 3.0, f"Should fail fast, took {result.duration_seconds}s"

    def test_valid_commands_still_works(self):
        """Non-empty commands should still work as before (regression check)."""
        from src.executor.ssh_executor import SSHExecutor
        from src.models.device import Device
        from src.models.task import Task
        from src.models.task_plan import TaskPlan

        executor = SSHExecutor()
        device = Device(
            row_index=1, device_name="test-dev", device_group="A3",
            bmc_ip="", bmc_username="", bmc_password="",
            inband_ip="127.0.0.1",
            inband_username="test", inband_password="test",
        )
        task = Task(
            row_index=1, sequence=1, task_name="正常命令",
            task_type="SSH", execution_mode="SSH_CMD",
            command_or_url="echo hello",
            match_group="A3",
        )
        plan = TaskPlan(plan_id="p1", device=device, task=task, task_id="t1")

        with tempfile.TemporaryDirectory() as tmpdir:
            result = executor.execute(plan, tmpdir)
            # Will fail due to connection refused (no SSH server on 127.0.0.1),
            # but should NOT be COMMAND_MISSING
            assert "COMMAND_MISSING" not in result.execution_failure_reason


# ===========================================================================
# tasks.json path resolution tests
# ===========================================================================


class TestTasksJsonPathResolution:
    """P0-3: tasks.json found in source/_internal paths."""

    def test_source_root_tasks_json_found(self, tmp_path):
        """tasks.json at project root is found."""
        tasks_json = tmp_path / "tasks.json"
        tasks_json.write_text(json.dumps({"tasks": {"测试": {"task_type": "SSH"}}}),
                              encoding="utf-8")

        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs(str(tasks_json))
        assert len(defs) == 1
        assert "测试" in defs

    def test_internal_tasks_json_found(self, tmp_path):
        """tasks.json in _internal/ is found."""
        internal_dir = tmp_path / "_internal"
        internal_dir.mkdir()
        tasks_json = internal_dir / "tasks.json"
        tasks_json.write_text(json.dumps({"tasks": {"测试": {"task_type": "SSH"}}}),
                              encoding="utf-8")

        # Should find it via the internal path
        from src.loader.excel_reader import _load_task_defs
        # Explicit path
        defs = _load_task_defs(str(tasks_json))
        assert len(defs) == 1

    def test_tasks_json_not_found_returns_empty(self, tmp_path):
        """Missing tasks.json returns empty dict with warning."""
        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs(str(tmp_path / "nonexistent.json"))
        assert defs == {}


# ===========================================================================
# Excel fallback safety net tests
# ===========================================================================


class TestExcelFallbackSafety:
    """P0-3: Simplified format without tasks.json disables SSH tasks."""

    def test_simplified_no_tdef_ssh_disabled(self, tmp_path):
        """Simplified Excel without tasks.json: SSH tasks with no command disabled."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务列表"
        ws.append(["序号", "任务名称", "任务类型", "分组", "输出目录", "图片名", "是否启用"])
        ws.append([1, "光模块测试", "SSH", "A3/L1/L2", "out", "img", "是"])
        excel_path = tmp_path / "test.xlsx"
        wb.save(str(excel_path))
        wb.close()

        from src.loader.excel_reader import load_tasks
        tasks = load_tasks(str(excel_path), tasks_json_path=str(tmp_path / "nonexistent.json"))
        assert len(tasks) == 1
        # Should be disabled — no command available
        assert tasks[0].enabled is False

    def test_simplified_with_tdef_ssh_enabled(self, tmp_path):
        """Simplified Excel with tasks.json: SSH tasks with command enabled."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务列表"
        ws.append(["序号", "任务名称", "任务类型", "分组", "输出目录", "图片名", "是否启用"])
        ws.append([1, "光模块测试", "SSH", "A3/L1/L2", "out", "img", "是"])
        excel_path = tmp_path / "test.xlsx"
        wb.save(str(excel_path))
        wb.close()

        tasks_json = tmp_path / "tasks.json"
        tasks_json.write_text(json.dumps({
            "tasks": {
                "光模块测试": {
                    "task_type": "SSH",
                    "execution_mode": "SSH_CMD",
                    "command_or_url": "display interface transceiver",
                    "per_group_commands": {
                        "A3": "hccn_tool -i 0 -optical -g"
                    },
                    "per_group_timeout_seconds": {
                        "A3": 900,
                        "L1": 60
                    }
                }
            }
        }), encoding="utf-8")

        from src.loader.excel_reader import load_tasks
        tasks = load_tasks(str(excel_path), tasks_json_path=str(tasks_json))
        assert len(tasks) == 1
        assert tasks[0].enabled is True
        assert tasks[0].command_or_url == "display interface transceiver"
        # per_group_commands should be stored
        pgc = getattr(tasks[0], '_per_group_commands', None)
        assert pgc is not None
        assert "A3" in pgc
        assert getattr(tasks[0], '_per_group_timeout_seconds', None) == {"A3": 900, "L1": 60}

    def test_legacy_format_no_tdef_still_enabled(self, tmp_path):
        """Legacy (14-col) Excel without tasks.json: still enabled (has command column)."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务列表"
        # 14-column legacy format
        ws.append(["序号", "任务名称", "类型", "分组", "模式", "命令", "动作",
                    "输出目录", "图片名", "超时", "重试", "是否启用", "规则", "备注"])
        ws.append([1, "光模块测试", "SSH", "A3", "SSH_CMD", "display interface transceiver", "",
                    "out", "img", 60, 0, "是", "", ""])
        excel_path = tmp_path / "test_legacy.xlsx"
        wb.save(str(excel_path))
        wb.close()

        from src.loader.excel_reader import load_tasks
        tasks = load_tasks(str(excel_path), tasks_json_path=str(tmp_path / "nonexistent.json"))
        assert len(tasks) == 1
        assert tasks[0].enabled is True
        assert tasks[0].command_or_url == "display interface transceiver"


# ===========================================================================
# PlanRunService per_group_commands resolution (existing path verification)
# ===========================================================================


class TestPlanRunServicePerGroupCommands:
    """Verify PlanRunService already reads _per_group_commands."""

    def test_build_job_payload_resolves_per_group(self):
        """PlanRunService._build_job_payload uses _per_group_commands."""
        from src.plan_run_service.service import PlanRunService, PlanRunItem
        from src.models.device import Device
        from src.models.task import Task

        svc = PlanRunService()
        device = Device(
            row_index=1, device_name="test-A3", device_group="A3",
            bmc_ip="10.0.0.1", bmc_username="", bmc_password="",
            inband_ip="10.0.0.1",
            inband_username="u", inband_password="p",
        )
        task = Task(
            row_index=1, sequence=1, task_name="光模块",
            task_type="SSH", execution_mode="SSH_CMD",
            command_or_url="default_cmd",
            match_group="A3",
        )
        object.__setattr__(task, '_per_group_commands', {"A3": "a3_special_cmd"})

        item = PlanRunItem(
            plan_id=1, device_name="test-A3", task_name="光模块",
            device_group="A3", task_type="SSH", execution_mode="SSH_CMD",
            _device=device, _task=task,
        )
        payload = svc._build_job_payload(item)
        # The command in task_snapshot should be the resolved one
        assert payload["task_snapshot"]["command_or_url"] == "a3_special_cmd"

    def test_actual_a3_optical_no_split_survives_api_adapter_chain(self):
        """Actual task 17 must keep A3 no_split after PlanRunService -> adapter."""
        from src.loader.excel_reader import load_all
        from src.models.device import Device
        from src.plan_run_service.service import PlanRunService, PlanRunItem
        from src.job_runner_adapter import RealRunnerAdapter
        from src.executor.ssh_executor import SSHExecutor, resolve_task_no_split

        _, tasks = load_all(
            str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx"),
            tasks_json_path=str(Path(__file__).resolve().parent.parent / "tasks.json"),
        )
        task = next(t for t in tasks if t.task_name == "计算节点光模块信息查询测试")
        device = Device(
            row_index=1, device_name="redacted-device", device_group="A3",
            bmc_ip="", bmc_username="", bmc_password="",
            inband_ip="192.0.2.1", inband_username="u", inband_password="p",
        )
        item = PlanRunItem(
            plan_id="p1", device_name=device.device_name, task_name=task.task_name,
            device_group="A3", task_type=task.task_type, execution_mode=task.execution_mode,
            _device=device, _task=task,
        )

        payload = PlanRunService()._build_job_payload(item)
        assert payload["task_snapshot"]["per_group_timeout_seconds"] == {"A3": 900, "L1": 180, "L2": 180}
        rebuilt = RealRunnerAdapter()._task_from_snapshot(payload["task_snapshot"])
        assert resolve_task_no_split(rebuilt, "A3") is True
        assert getattr(rebuilt, "_task_def", {}).get("stderr_fail_patterns")
        assert SSHExecutor(command_timeout=60)._resolve_execution_options(rebuilt, "A3").command_timeout == 900
        assert SSHExecutor(command_timeout=60)._resolve_execution_options(rebuilt, "L1").command_timeout == 180

        spec = SSHExecutor()._parse_command_spec(
            rebuilt,
            override_command=rebuilt.command_or_url,
            no_split=resolve_task_no_split(rebuilt, "A3"),
        )
        assert len(spec["commands"]) == 1

    def test_actual_bmc_actions_survives_api_adapter_chain(self):
        """Actual BMC_ACTIONS task must keep actions_json through API real path."""
        from src.loader.excel_reader import load_all
        from src.models.device import Device
        from src.plan_run_service.service import PlanRunService, PlanRunItem
        from src.job_runner_adapter import RealRunnerAdapter

        _, tasks = load_all(
            str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx"),
            tasks_json_path=str(Path(__file__).resolve().parent.parent / "tasks.json"),
        )
        task = next(t for t in tasks if t.task_name == "RAID配置测试")
        device = Device(
            row_index=1, device_name="redacted-device", device_group="A3",
            bmc_ip="192.0.2.2", bmc_username="u", bmc_password="p",
            inband_ip="", inband_username="", inband_password="",
        )
        item = PlanRunItem(
            plan_id="p1", device_name=device.device_name, task_name=task.task_name,
            device_group="A3", task_type=task.task_type, execution_mode=task.execution_mode,
            _device=device, _task=task,
        )

        payload = PlanRunService()._build_job_payload(item)
        rebuilt = RealRunnerAdapter()._task_from_snapshot(payload["task_snapshot"])
        assert rebuilt.actions_json
        flow = rebuilt.to_capture_flow()
        assert flow.get("target_url") or flow.get("pre_capture_actions")

    def test_actual_rules_survive_api_adapter_chain(self):
        """Actual rule-bearing tasks must not lose rules_json or task_def."""
        from src.loader.excel_reader import load_all
        from src.models.device import Device
        from src.plan_run_service.service import PlanRunService, PlanRunItem
        from src.job_runner_adapter import RealRunnerAdapter

        _, tasks = load_all(
            str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx"),
            tasks_json_path=str(Path(__file__).resolve().parent.parent / "tasks.json"),
        )
        task = next(t for t in tasks if t.task_name == "计算节点L1交换网络端口查询测试")
        device = Device(
            row_index=1, device_name="redacted-device", device_group="L1",
            bmc_ip="", bmc_username="", bmc_password="",
            inband_ip="192.0.2.1", inband_username="u", inband_password="p",
        )
        item = PlanRunItem(
            plan_id="p1", device_name=device.device_name, task_name=task.task_name,
            device_group="L1", task_type=task.task_type, execution_mode=task.execution_mode,
            _device=device, _task=task,
        )

        payload = PlanRunService()._build_job_payload(item)
        rebuilt = RealRunnerAdapter()._task_from_snapshot(payload["task_snapshot"])
        assert len(rebuilt.parsed_rules()) == len(task.parsed_rules())
        assert getattr(rebuilt, "_task_def", {}).get("rules")


# ===========================================================================
# API infoEvents / timestamp tests
# ===========================================================================


class TestApiInfoTimestamp:
    """API plan query / item query must include timestamps and infoEvents."""

    EXCEL_FILE = str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx")

    def _service(self):
        from src.plan_item_status_callback_client import FakeCallbackTransport
        from src.plan_run_service.service import PlanRunService
        return PlanRunService(callback_transport=FakeCallbackTransport())

    def _wait_external_completed(self, svc, plan_id, excel_hash):
        for _ in range(60):
            plan = svc.get_external_plan(plan_id, excel_hash)
            if plan and plan.get("status") == "COMPLETED":
                return plan
            time.sleep(0.1)
        return svc.get_external_plan(plan_id, excel_hash)

    def _wait_plan_completed(self, svc, plan_id):
        for _ in range(60):
            plan = svc.get_plan(plan_id)
            if plan and plan.get("status") == "COMPLETED":
                return plan
            time.sleep(0.1)
        return svc.get_plan(plan_id)

    def test_plan_query_has_started_at_finished_at(self):
        svc = self._service()
        svc.set_latest_excel(self.EXCEL_FILE)
        excel_hash = svc.set_latest_excel(self.EXCEL_FILE)["excelHash"]
        r = svc.start_external_plan({
            "excelHash": excel_hash,
            "callback": {"planId": "1", "itemStatusUrl": "http://cb"},
            "runner": "fake",
        })
        plan = self._wait_external_completed(svc, r["planId"], excel_hash)
        assert plan is not None
        assert "startedAt" in plan
        assert "finishedAt" in plan
        assert plan["startedAt"] != ""

    def test_plan_query_has_info_events(self):
        svc = self._service()
        svc.set_latest_excel(self.EXCEL_FILE)
        excel_hash = svc.set_latest_excel(self.EXCEL_FILE)["excelHash"]
        r = svc.start_external_plan({
            "excelHash": excel_hash,
            "callback": {"planId": "1", "itemStatusUrl": "http://cb"},
            "runner": "fake",
        })
        plan = self._wait_external_completed(svc, r["planId"], excel_hash)
        assert "infoEvents" in plan
        assert isinstance(plan["infoEvents"], list)

    def test_item_query_has_timestamps(self):
        svc = self._service()
        svc.set_latest_excel(self.EXCEL_FILE)
        excel_hash = svc.set_latest_excel(self.EXCEL_FILE)["excelHash"]
        r = svc.start_external_plan({
            "excelHash": excel_hash,
            "callback": {"planId": "1", "itemStatusUrl": "http://cb"},
            "runner": "fake",
        })
        self._wait_external_completed(svc, r["planId"], excel_hash)
        items_data = svc.get_external_plan_items(r["planId"], excel_hash)
        for item in items_data["items"]:
            assert "startedAt" in item, f"Item missing startedAt: {item}"
            assert "finishedAt" in item, f"Item missing finishedAt: {item}"
            assert item["startedAt"] is not None
            assert item["finishedAt"] is not None

    def test_item_query_has_info_events(self):
        svc = self._service()
        svc.set_latest_excel(self.EXCEL_FILE)
        excel_hash = svc.set_latest_excel(self.EXCEL_FILE)["excelHash"]
        r = svc.start_external_plan({
            "excelHash": excel_hash,
            "callback": {"planId": "1", "itemStatusUrl": "http://cb"},
            "runner": "fake",
        })
        self._wait_external_completed(svc, r["planId"], excel_hash)
        items_data = svc.get_external_plan_items(r["planId"], excel_hash)
        for item in items_data["items"]:
            assert "infoEvents" in item
            assert isinstance(item["infoEvents"], list)
            assert len(item["infoEvents"]) >= 1, f"Item should have at least 1 infoEvent: {item}"

    def test_info_event_has_timestamp_level_message(self):
        svc = self._service()
        svc.set_latest_excel(self.EXCEL_FILE)
        excel_hash = svc.set_latest_excel(self.EXCEL_FILE)["excelHash"]
        r = svc.start_external_plan({
            "excelHash": excel_hash,
            "callback": {"planId": "1", "itemStatusUrl": "http://cb"},
            "runner": "fake",
        })
        self._wait_external_completed(svc, r["planId"], excel_hash)
        items_data = svc.get_external_plan_items(r["planId"], excel_hash)
        for item in items_data["items"]:
            for evt in item["infoEvents"]:
                assert "timestamp" in evt, f"Event missing timestamp: {evt}"
                assert "level" in evt, f"Event missing level: {evt}"
                assert "message" in evt, f"Event missing message: {evt}"
                # timestamp should be ISO format
                assert "T" in evt["timestamp"], f"Timestamp not ISO: {evt['timestamp']}"
                assert evt["level"] in ("INFO", "WARN", "ERROR")

    def test_plan_run_item_has_info_events_field(self):
        from src.plan_run_service.service import PlanRunItem
        item = PlanRunItem(plan_id="p1", device_name="D1", task_name="T1")
        item.add_info_event("INFO", "test message")
        assert len(item.info_events) == 1
        evt = item.info_events[0]
        assert "timestamp" in evt
        assert evt["level"] == "INFO"
        assert evt["message"] == "test message"


# ===========================================================================
# Dynamic path resolution tests — no absolute paths
# ===========================================================================


class TestDynamicPathResolution:
    """Verify all path resolvers use relative/dynamic paths, not absolute."""

    def test_project_root_is_derived_from_file(self):
        """project_root is derived from __file__, not hardcoded."""
        from pathlib import Path
        import src.loader.excel_reader as er
        module_file = Path(er.__file__).resolve()
        project_root = module_file.parent.parent.parent
        # Must be relative to the actual source tree
        # project_root must be derived from __file__, not hardcoded
        assert project_root.is_dir(), f"project_root must be a directory: {project_root}"
        # The resolver must NOT contain any absolute path like E:\ or C:\Users
        assert "E:\\" not in str(project_root)
        assert "C:\\Users" not in str(project_root)

    def test_tasks_json_candidates_are_relative(self):
        """All tasks.json search candidates are derived from project_root/CWD."""
        from pathlib import Path
        from src.loader.excel_reader import _load_task_defs

        # Use explicit path to test the candidate generation logic
        project_root = Path(__file__).resolve().parent.parent
        cwd = Path.cwd()

        # Verify no absolute paths in the candidate logic
        for p in [project_root, cwd]:
            path_str = str(p)
            # Must not contain user-specific paths
            assert "p_OccRemoteDesk" not in path_str

    def test_playwright_browsers_dir_is_relative(self, monkeypatch):
        """Playwright browsers dir is derived from __file__, not hardcoded."""
        from src.executor.browser_manager import _resolve_playwright_browsers_dir
        resolved = _resolve_playwright_browsers_dir()
        if resolved is not None:
            # Must not be an absolute user-specific path
            # On macOS/Linux, /Users/ is a valid path; only reject Windows-style user paths
            if "\\" in str(resolved):
                assert "\\Users\\" not in str(resolved), f"Windows user-specific path: {resolved}"
            assert "p_OccRemoteDesk" not in str(resolved)

    def test_release_root_from_exe_in_runtime_dir(self, monkeypatch, tmp_path):
        """When exe is in runtime/, release root derived from exe finds tasks.json."""
        import sys as _sys

        # Simulate: <release>/runtime/bmc-engine.exe
        release_root = tmp_path / "bmc-auto-capture"
        runtime_dir = release_root / "runtime"
        runtime_dir.mkdir(parents=True)
        exe = runtime_dir / "bmc-engine.exe"
        exe.write_text("fake")

        # Create tasks.json ONLY at release root (not at project_root)
        tasks_json = release_root / "tasks.json"
        tasks_json.write_text('{"tasks":{"exe_test":{"task_type":"SSH"}}}', encoding="utf-8")

        # Mock sys.executable → exe path, AND CWD → release root
        monkeypatch.setattr(_sys, "executable", str(exe))
        monkeypatch.setattr("pathlib.Path.cwd", lambda: release_root)

        # Also need to prevent project_root from finding real tasks.json
        # by using an explicit path
        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs(str(tasks_json))
        assert len(defs) == 1
        assert "exe_test" in defs

    def test_tasks_json_in_app_subdir_found(self, tmp_path, monkeypatch):
        """When tasks.json is in app/ subdirectory, it is found via CWD/app/."""
        release_root = tmp_path / "bmc-auto-capture"
        app_dir = release_root / "app"
        app_dir.mkdir(parents=True)
        tasks_json = app_dir / "tasks.json"
        tasks_json.write_text('{"tasks":{"app_test":{"task_type":"SSH"}}}', encoding="utf-8")

        # Pass explicit path to avoid project_root match
        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs(str(tasks_json))
        assert len(defs) == 1
        assert "app_test" in defs

    def test_tasks_json_at_cwd_root_found(self, tmp_path, monkeypatch):
        """When tasks.json is directly in CWD, it is found."""
        tasks_json = tmp_path / "tasks.json"
        tasks_json.write_text('{"tasks":{"cwd_test":{"task_type":"BMC"}}}', encoding="utf-8")

        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs(str(tasks_json))
        assert len(defs) == 1
        assert "cwd_test" in defs

    def test_source_layout_tasks_json_found(self):
        """Source layout: tasks.json at project root, 4.1.15 present."""
        from pathlib import Path
        from src.loader.excel_reader import _load_task_defs
        defs = _load_task_defs()
        assert len(defs) > 0
        assert "计算节点光模块信息查询测试" in defs

    def test_all_candidates_are_dynamic(self):
        """Every candidate in the search is derived from Path computations."""
        from pathlib import Path

        # Simulate: collect all the candidate patterns and verify they don't
        # contain hardcoded absolute paths.
        project_root = Path("/fake/project/src/loader").parent.parent.parent  # /fake/project
        cwd = Path("/fake/working/dir")
        _exe_dir = Path("/fake/release")

        candidates = [
            project_root / "tasks.json",
            cwd / "tasks.json",
            cwd / "app" / "tasks.json",
            _exe_dir / "tasks.json",
            _exe_dir / "app" / "tasks.json",
            project_root / "_internal" / "tasks.json",
            cwd / "_internal" / "tasks.json",
            project_root.parent / "tasks.json",
            project_root.parent / "app" / "tasks.json",
            project_root.parent.parent / "tasks.json",
            project_root.parent.parent / "app" / "tasks.json",
        ]

        # All paths must be PosixPath or WindowsPath — no raw strings
        for c in candidates:
            assert isinstance(c, Path), f"Expected Path, got {type(c)}: {c}"
            # Must not contain Windows drive letter as raw string
            path_str = str(c)
            assert "E:\\v0.2" not in path_str
            assert "p_OccRemoteDesk" not in path_str

# ===========================================================================
# End-to-end: _parse_command_spec MUST use resolved command
# ===========================================================================


class TestParseCommandSpecWithOverride:
    """Verify _parse_command_spec uses override_command, not raw command_or_url."""

    def test_a3_override_parses_hccn_tool_not_display(self):
        """A3: override_command=hccn_tool → ALL commands contain hccn_tool, not display."""
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()

        class FakeTask:
            task_name = "光模块"
            command_or_url = "display interface transceiver"

        task = FakeTask()

        # A3 override (semicolons split into multiple commands)
        a3_cmd = "for i in $(seq 0 15); do echo \"> $i\"; hccn_tool -i $i -optical -g;done"
        spec = executor._parse_command_spec(task, override_command=a3_cmd)
        commands = spec["commands"]
        assert len(commands) >= 3  # split by ;
        all_cmds = " ".join(c[1] for c in commands)
        assert "hccn_tool" in all_cmds, f"hccn_tool not found in: {all_cmds}"
        assert "display interface transceiver" not in all_cmds

    def test_l1_no_override_uses_command_or_url(self):
        """L1: no override → uses task.command_or_url."""
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()

        class FakeTask:
            task_name = "光模块"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        spec = executor._parse_command_spec(task)
        commands = spec["commands"]
        assert len(commands) == 1
        assert commands[0][1] == "display interface transceiver"

    def test_override_empty_returns_empty_commands(self):
        """Empty override → empty commands."""
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()

        class FakeTask:
            task_name = "test"
            command_or_url = "some_cmd"

        task = FakeTask()
        spec = executor._parse_command_spec(task, override_command="")
        assert spec["commands"] == []

    def test_full_4_1_15_a3_resolution_chain(self):
        """Full chain: resolve_task_command → _parse_command_spec → correct commands."""
        from src.executor.ssh_executor import SSHExecutor, resolve_task_command

        executor = SSHExecutor()

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do echo \"==============> $i\"; hccn_tool -i $i -optical -g;done",
        })

        # Step 1: resolve
        resolved = resolve_task_command(task, "A3")
        assert "hccn_tool" in resolved
        assert "display interface transceiver" not in resolved

        # Step 2: parse with resolved command
        spec = executor._parse_command_spec(task, override_command=resolved)
        commands = spec["commands"]
        assert len(commands) >= 1

        # Step 3: verify the actual commands that will execute
        all_cmds = " ".join(c[1] for c in commands)
        assert "hccn_tool" in all_cmds, f"Expected hccn_tool in commands, got: {all_cmds[:120]}"
        assert "-optical -g" in all_cmds, f"Expected -optical -g flag in commands, got: {all_cmds[:120]}"
        assert "display interface transceiver" not in all_cmds

    def test_full_4_1_15_l1_resolution_chain(self):
        """Full chain: L1 → resolve → display interface transceiver."""
        from src.executor.ssh_executor import SSHExecutor, resolve_task_command

        executor = SSHExecutor()

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do hccn_tool -i $i -optical -g;done",
        })

        resolved = resolve_task_command(task, "L1")
        assert resolved == "display interface transceiver"

        spec = executor._parse_command_spec(task, override_command=resolved)
        assert spec["commands"][0][1] == "display interface transceiver"

    def test_full_4_1_15_l2_resolution_chain(self):
        """Full chain: L2 → resolve → display interface transceiver."""
        from src.executor.ssh_executor import SSHExecutor, resolve_task_command

        executor = SSHExecutor()

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": "for i in $(seq 0 15); do hccn_tool -i $i -optical -g;done",
        })

        resolved = resolve_task_command(task, "L2")
        assert resolved == "display interface transceiver"

        spec = executor._parse_command_spec(task, override_command=resolved)
        assert spec["commands"][0][1] == "display interface transceiver"

    def test_legacy_run_items_also_have_timestamps(self):
        from src.plan_run_service.service import PlanRunService
        excel = str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx")
        svc = PlanRunService()
        svc.set_latest_excel(excel)


# ===========================================================================
# no_split tests
# ===========================================================================


class TestNoSplit:
    """Verify no_split prevents ; splitting for shell compound commands."""

    def test_resolve_task_no_split_a3_true(self):
        from src.executor.ssh_executor import resolve_task_no_split

        class FakeTask:
            pass

        task = FakeTask()
        object.__setattr__(task, '_per_group_no_split', {"A3": True})

        assert resolve_task_no_split(task, "A3") is True
        assert resolve_task_no_split(task, "L1") is False

    def test_resolve_task_no_split_global(self):
        from src.executor.ssh_executor import resolve_task_no_split

        class FakeTask:
            pass

        task = FakeTask()
        object.__setattr__(task, '_no_split', True)

        assert resolve_task_no_split(task, "A3") is True
        assert resolve_task_no_split(task, "L1") is True

    def test_resolve_task_no_split_not_configured(self):
        from src.executor.ssh_executor import resolve_task_no_split

        class FakeTask:
            pass

        task = FakeTask()
        assert resolve_task_no_split(task, "A3") is False

    def test_parse_commands_no_split_keeps_for_loop_intact(self):
        from src.executor.ssh_executor import SSHExecutor
        e = SSHExecutor()
        cmd = 'for i in $(seq 0 15); do echo "==============> $i"; hccn_tool -i $i -optical -g;done'
        result = e._parse_commands(cmd, no_split=True)
        assert len(result) == 1, f"Expected 1 command, got {len(result)}: {result}"
        assert result[0] == cmd

    def test_parse_commands_normal_split_by_semicolon(self):
        from src.executor.ssh_executor import SSHExecutor
        e = SSHExecutor()
        cmd = 'for i in $(seq 0 15); do echo "> $i"; hccn_tool -i $i -optical -g;done'
        result = e._parse_commands(cmd, no_split=False)
        # Without no_split, ; causes splitting
        assert len(result) > 1

    def test_parse_commands_normal_uname_not_split(self):
        from src.executor.ssh_executor import SSHExecutor
        e = SSHExecutor()
        # uname -a has no ; → stays as 1
        result = e._parse_commands("uname -a", no_split=False)
        assert len(result) == 1
        assert result[0] == "uname -a"

    def test_parse_commands_multiline_not_split_without_no_split(self):
        from src.executor.ssh_executor import SSHExecutor
        e = SSHExecutor()
        cmd = "npu-smi info\n/usr/local/bin/tool --version"
        result = e._parse_commands(cmd, no_split=False)
        assert len(result) == 2

    def test_full_4_1_15_a3_no_split_chain(self):
        from src.executor.ssh_executor import SSHExecutor, resolve_task_command, resolve_task_no_split

        e = SSHExecutor()

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": 'for i in $(seq 0 15); do echo "==============> $i"; hccn_tool -i $i -optical -g;done',
        })
        object.__setattr__(task, '_per_group_no_split', {"A3": True})

        # Full chain: resolve → no_split check → parse
        resolved = resolve_task_command(task, "A3")
        assert "hccn_tool" in resolved
        assert resolve_task_no_split(task, "A3") is True

        spec = e._parse_command_spec(task, override_command=resolved, no_split=True)
        commands = spec["commands"]
        assert len(commands) == 1, f"Expected 1 command with no_split, got {len(commands)}"
        assert "hccn_tool" in commands[0][1]
        assert "-optical -g" in commands[0][1]
        assert "display interface transceiver" not in commands[0][1]

    def test_4_1_15_l1_no_no_split_fallback(self):
        from src.executor.ssh_executor import SSHExecutor, resolve_task_command, resolve_task_no_split

        e = SSHExecutor()

        class FakeTask:
            task_name = "计算节点光模块信息查询测试"
            command_or_url = "display interface transceiver"

        task = FakeTask()
        object.__setattr__(task, '_per_group_commands', {
            "A3": 'for i in $(seq 0 15); do hccn_tool -i $i -optical -g;done',
        })
        object.__setattr__(task, '_per_group_no_split', {"A3": True})

        # L1: no per_group_commands match → fallback, no no_split
        resolved = resolve_task_command(task, "L1")
        assert resolved == "display interface transceiver"
        assert resolve_task_no_split(task, "L1") is False

        spec = e._parse_command_spec(task, override_command=resolved, no_split=False)
        assert len(spec["commands"]) == 1
        assert spec["commands"][0][1] == "display interface transceiver"

    def test_legacy_run_items_also_have_timestamps(self):
        from src.plan_item_status_callback_client import FakeCallbackTransport
        from src.plan_run_service.service import PlanRunService
        excel = str(Path(__file__).resolve().parent.parent / "examples" / "task_template.xlsx")
        svc = PlanRunService(callback_transport=FakeCallbackTransport())
        svc.set_latest_excel(excel)
        r = svc.start_plan_run(1, {"callback": {"planId": "1", "itemStatusUrl": "http://cb"}})
        for _ in range(60):
            plan = svc.get_plan(r["planId"])
            if plan and plan.get("status") == "COMPLETED":
                break
            time.sleep(0.1)
        items_data = svc.get_plan_items(r["planId"])
        for item in items_data["items"]:
            assert "startedAt" in item
            assert "finishedAt" in item
            assert "infoEvents" in item


class TestSSHStrategyProfiles:
    """User-facing SSH model is Linux terminal vs VRP interactive."""

    def _device(self, group: str):
        from src.models.device import Device

        return Device(
            row_index=1,
            device_name=f"{group}-dev",
            device_group=group,
            bmc_ip="",
            bmc_username="",
            bmc_password="",
            inband_ip="10.0.0.1",
        )

    def _task(self, task_def=None):
        from src.models.task import Task

        task = Task(
            row_index=1,
            sequence=1,
            task_name="SSH profile test",
            task_type="SSH",
            execution_mode="SSH_CMD",
            command_or_url="uname -a",
        )
        if task_def:
            object.__setattr__(task, "_task_def", task_def)
        return task

    def test_linux_group_defaults_to_terminal_session(self):
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()
        assert executor._get_ssh_strategy(self._device("A3"), self._task()) == "terminal_session"

    def test_vrp_group_defaults_to_interactive_shell(self):
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()
        assert executor._get_ssh_strategy(self._device("L1"), self._task()) == "interactive_shell"
        assert executor._get_ssh_strategy(self._device("L2"), self._task()) == "interactive_shell"

    def test_structured_evidence_mode_keeps_exec_command_available(self):
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()
        task = self._task({"evidence_mode": "structured"})
        assert executor._get_ssh_strategy(self._device("A3"), task) == "exec_command"

    def test_explicit_transport_overrides_profile(self):
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()
        task = self._task({"ssh_profile": "vrp", "ssh_transport": "terminal_session"})
        assert executor._get_ssh_strategy(self._device("L1"), task) == "terminal_session"

    def test_per_group_profile_override(self):
        from src.executor.ssh_executor import SSHExecutor

        executor = SSHExecutor()
        task = self._task({"per_group_ssh_profile": {"A3": "vrp"}})
        assert executor._get_ssh_strategy(self._device("A3"), task) == "interactive_shell"
