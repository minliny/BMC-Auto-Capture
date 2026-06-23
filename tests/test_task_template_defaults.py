from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
TASK_TEMPLATE = PROJECT_ROOT / "examples" / "task_template.xlsx"
DEFAULT_OUTPUT_DIR_TEMPLATE = "{任务序号}_{任务名称}/{设备分类}"


def test_task_template_uses_underscore_output_dir_default():
    wb = load_workbook(TASK_TEMPLATE, read_only=True, data_only=True)
    try:
        tasks = wb["任务列表"]
        output_dir_values = [
            row[4]
            for row in tasks.iter_rows(min_row=2, values_only=True)
            if any(cell is not None for cell in row)
        ]
        assert output_dir_values
        assert set(output_dir_values) == {DEFAULT_OUTPUT_DIR_TEMPLATE}

        notes = wb["说明"]
        assert DEFAULT_OUTPUT_DIR_TEMPLATE in str(notes["A3"].value)
    finally:
        wb.close()
