"""
Excel V2 reader — parses 设备信息 and 任务列表 sheets into Device and Task lists.

Task execution details (URL, commands, rules) are resolved from tasks.json
by matching task_id first. task_name matching is retained only as a migration
fallback for older spreadsheets and task definitions.
"""

from __future__ import annotations
import json
import logging
import os
import re
from pathlib import Path
from typing import Any
import openpyxl

from ..models.device import Device
from ..models.task import Task

logger = logging.getLogger("bmc_auto_capture.loader")

# Simplified Excel columns for tasks (v2.1):
# 0:任务序号 1:任务名称 2:任务类型 3:设备分组 4:标签
# 5:输出目录模板 6:图片命名格式 7:是否启用
SIMPLIFIED_TASK_COLS = 7


def _str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _bool(val: Any) -> bool:
    s = _str(val).lower()
    if s in ("是", "yes", "true", "1", "y", "启用"):
        return True
    if s in ("否", "no", "false", "0", "n", "禁用", ""):
        return False
    return bool(val)


def _int(val: Any, default: int = 0) -> int:
    """Parse integer. Supports dotted notation like '4.1.7' → 4001007."""
    try:
        return int(val)
    except (ValueError, TypeError):
        pass
    # Try dotted format: 4.1.7 → 4*1000000 + 1*1000 + 7 = 4001007
    s = str(val).strip()
    if s and all(p.isdigit() for p in s.split(".") if p):
        parts = [int(p) for p in s.split(".")]
        result = 0
        for p in parts:
            result = result * 1000 + p
        return result
    return default


def _parse_tags(raw: str) -> tuple[str, ...]:
    if not raw:
        return ()
    parts = [t.strip() for t in raw.replace("，", ",").split(",") if t.strip()]
    return tuple(parts)


# Column name → canonical field name mapping (header-based, order-independent)
DEVICE_HEADER_MAP: dict[str, str] = {}
TASK_HEADER_MAP: dict[str, str] = {}

def _parse_bilingual_header(raw: str) -> list[str]:
    """Parse a bilingual header like '设备名称(DeviceName)' into [中文, 英文].

    Supports: '设备名称(DeviceName)' → ['设备名称', 'DeviceName']
              '带外管理IP(OOB_IP)'  → ['带外管理IP', 'OOB_IP']
              'DeviceName'          → ['DeviceName']
              '设备名称'            → ['设备名称']
    """
    raw = raw.strip()
    # Match "中文(English)" pattern
    m = re.match(r'^(.+?)\(([A-Za-z_][A-Za-z0-9_]*)\)$', raw)
    if m:
        return [m.group(1).strip(), m.group(2).strip()]
    return [raw]


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Build header_name → column_index map from the header row.

    Supports:
    - Chinese-only: 设备名称
    - Bilingual: 设备名称(DeviceName)
    - English-only: DeviceName

    Headers are matched by looking up canonical field names via DEVICE_HEADER_MAP.
    For bilingual headers, both the Chinese and English parts are mapped.
    """
    if not DEVICE_HEADER_MAP:
        _init_header_map()

    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        h_clean = _str(h)
        if not h_clean:
            continue

        # Parse bilingual: try each part
        parts = _parse_bilingual_header(h_clean)
        matched = False
        for part in parts:
            field = DEVICE_HEADER_MAP.get(part, "")
            if field:
                col_map[field] = idx
                matched = True
                break

        if matched:
            continue

        # Fallback: case-insensitive match against all known keys
        for key, val in DEVICE_HEADER_MAP.items():
            if key.lower() == h_clean.lower():
                col_map[val] = idx
                matched = True
                break

    return col_map


def _build_task_header_map(headers: list[str]) -> dict[str, int]:
    """Build canonical task field → column_index map from the task header row."""
    if not TASK_HEADER_MAP:
        _init_task_header_map()

    col_map: dict[str, int] = {}
    for idx, h in enumerate(headers):
        h_clean = _str(h)
        if not h_clean:
            continue
        parts = _parse_bilingual_header(h_clean)
        matched = False
        for part in parts:
            field = TASK_HEADER_MAP.get(part, "")
            if field:
                col_map[field] = idx
                matched = True
                break
        if matched:
            continue
        for key, val in TASK_HEADER_MAP.items():
            if key.lower() == h_clean.lower():
                col_map[val] = idx
                break
    return col_map


def _init_header_map():
    """One-time init of header name → field mapping."""
    # Bilingual entries: "中文(英文)" → canonical field name
    _bilingual: list[tuple[str, str]] = [
        ("设备名称(DeviceName)", "device_name"),
        ("设备分组(DeviceGroup)", "device_group"),
        ("设备分类(DeviceGroup)", "device_group"),
        ("带外管理IP(OOB_IP)", "bmc_ip"),
        ("带外管理用户名(OOB_Username)", "bmc_username"),
        ("带外管理密码(OOB_Password)", "bmc_password"),
        ("带内管理IP(IB_IP)", "inband_ip"),
        ("带内管理用户名(IB_Username)", "inband_username"),
        ("带内管理密码(IB_Password)", "inband_password"),
        ("设备是否启用(DeviceEnabled)", "enabled"),
        ("标签(Tags)", "tags"),
        ("是否全量截图(FullScreenshot)", "full_screenshot"),
        ("截图模式(ScreenshotMode)", "screenshot_mode"),
    ]
    for cn_en, field in _bilingual:
        DEVICE_HEADER_MAP[cn_en] = field

    DEVICE_HEADER_MAP.update({
        # device_group — accepts both "设备分组" and legacy "设备分类"
        "设备分组": "device_group",
        "设备分类": "device_group",
        "DeviceGroup": "device_group",
        "device_group": "device_group",
        # device_name
        "设备名称": "device_name",
        "DeviceName": "device_name",
        "device_name": "device_name",
        # bmc_ip
        "带外管理IP": "bmc_ip",
        "OOB_IP": "bmc_ip",
        "bmc_ip": "bmc_ip",
        "oob_ip": "bmc_ip",
        "BMC IP": "bmc_ip",
        "BMC IP地址": "bmc_ip",
        # enabled
        "设备是否启用": "enabled",
        "DeviceEnabled": "enabled",
        "是否启用": "enabled",
        "enabled": "enabled",
        # bmc_username
        "带外管理用户名": "bmc_username",
        "OOB_Username": "bmc_username",
        "bmc_username": "bmc_username",
        "BMC用户名": "bmc_username",
        "BMC账号": "bmc_username",
        # bmc_password
        "带外管理密码": "bmc_password",
        "OOB_Password": "bmc_password",
        "bmc_password": "bmc_password",
        "BMC密码": "bmc_password",
        # inband_ip
        "带内管理IP": "inband_ip",
        "IB_IP": "inband_ip",
        "inband_ip": "inband_ip",
        "ib_ip": "inband_ip",
        "SSH IP": "inband_ip",
        # inband_username
        "带内管理用户名": "inband_username",
        "IB_Username": "inband_username",
        "inband_username": "inband_username",
        "SSH用户名": "inband_username",
        "带内账号": "inband_username",
        # inband_password
        "带内管理密码": "inband_password",
        "IB_Password": "inband_password",
        "inband_password": "inband_password",
        "SSH密码": "inband_password",
        "带内密码": "inband_password",
        # tags
        "标签": "tags",
        "Tags": "tags",
        "tags": "tags",
    })


def _init_task_header_map():
    """One-time init of task header name → field mapping."""
    bilingual: list[tuple[str, str]] = [
        ("任务ID(TaskID)", "task_id"),
        ("任务序号(TaskSequence)", "sequence"),
        ("任务名称(TaskName)", "task_name"),
        ("任务类型(TaskType)", "task_type"),
        ("设备分组(DeviceGroup)", "match_group"),
        ("设备分类(DeviceGroup)", "match_group"),
        ("截图保存目录(OutputDir)", "output_dir_template"),
        ("输出目录(OutputDir)", "output_dir_template"),
        ("图片命名格式(FileNamePattern)", "image_name_template"),
        ("文件名模板(FileNamePattern)", "image_name_template"),
        ("是否启用(Enabled)", "enabled"),
        ("是否全量截图(FullScreenshot)", "full_screenshot"),
        ("截图模式(ScreenshotMode)", "screenshot_mode"),
        ("执行模式(ExecutionMode)", "execution_mode"),
        ("执行命令(CommandOrUrl)", "command_or_url"),
        ("动作JSON(ActionsJSON)", "actions_json"),
        ("超时时间(TimeoutSeconds)", "timeout_seconds"),
        ("重试次数(RetryCount)", "retry_count"),
        ("规则JSON(RulesJSON)", "rules_json"),
    ]
    for cn_en, field in bilingual:
        TASK_HEADER_MAP[cn_en] = field

    TASK_HEADER_MAP.update({
        "任务ID": "task_id",
        "TaskID": "task_id",
        "task_id": "task_id",
        "taskId": "task_id",
        "任务序号": "sequence",
        "序号": "sequence",
        "TaskSequence": "sequence",
        "sequence": "sequence",
        "任务名称": "task_name",
        "TaskName": "task_name",
        "task_name": "task_name",
        "任务类型": "task_type",
        "类型": "task_type",
        "TaskType": "task_type",
        "task_type": "task_type",
        "设备分组": "match_group",
        "设备分类": "match_group",
        "分组": "match_group",
        "DeviceGroup": "match_group",
        "match_group": "match_group",
        "截图保存目录": "output_dir_template",
        "输出目录": "output_dir_template",
        "OutputDir": "output_dir_template",
        "output_dir_template": "output_dir_template",
        "图片命名格式": "image_name_template",
        "文件名模板": "image_name_template",
        "图片名": "image_name_template",
        "FileNamePattern": "image_name_template",
        "image_name_template": "image_name_template",
        "是否启用": "enabled",
        "Enabled": "enabled",
        "enabled": "enabled",
        "是否全量截图": "full_screenshot",
        "FullScreenshot": "full_screenshot",
        "full_screenshot": "full_screenshot",
        "截图模式": "screenshot_mode",
        "ScreenshotMode": "screenshot_mode",
        "screenshot_mode": "screenshot_mode",
        "执行模式": "execution_mode",
        "模式": "execution_mode",
        "ExecutionMode": "execution_mode",
        "execution_mode": "execution_mode",
        "执行命令": "command_or_url",
        "命令": "command_or_url",
        "CommandOrUrl": "command_or_url",
        "command_or_url": "command_or_url",
        "URL": "command_or_url",
        "SSH_CMD": "command_or_url",
        "动作JSON": "actions_json",
        "动作": "actions_json",
        "ActionsJSON": "actions_json",
        "actions_json": "actions_json",
        "超时时间": "timeout_seconds",
        "超时": "timeout_seconds",
        "TimeoutSeconds": "timeout_seconds",
        "timeout_seconds": "timeout_seconds",
        "重试次数": "retry_count",
        "重试": "retry_count",
        "RetryCount": "retry_count",
        "retry_count": "retry_count",
        "规则JSON": "rules_json",
        "规则": "rules_json",
        "RulesJSON": "rules_json",
        "rules_json": "rules_json",
    })


def load_devices(filepath: str | Path, sheet_name: str = "设备信息") -> list[Device]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]

    # Read header row (row 1)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_str(v) for v in header_row] if header_row else []
    col_map = _build_header_map(headers)

    if not col_map:
        wb.close()
        raise ValueError(
            f"No recognized headers in '{sheet_name}'. "
            f"Found: {headers}. Expected headers like 设备名称, 带外管理IP, etc."
        )

    logger.info("Device sheet headers: %s → mapped fields: %s",
                headers, list(col_map.keys()))

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    def _get(field: str) -> str:
        idx = col_map.get(field, -1)
        if idx >= 0 and idx < len(vals):
            return vals[idx]
        return ""

    devices: list[Device] = []
    for i, row in enumerate(rows):
        if all(v is None for v in row):
            continue

        vals = [_str(v) for v in row]
        tags = _parse_tags(_get("tags"))
        tags_str = ", ".join(tags) if tags else ""

        # Validate BMC IP: reject obvious non-IP values
        bmc_raw = _get("bmc_ip")
        enabled_raw = _get("enabled")
        if bmc_raw in ("是", "否", "启用", "禁用", "yes", "no", "true", "false"):
            logger.warning(
                "Row %d: bmc_ip='%s' looks like an enabled flag — "
                "possible column misalignment. Check Excel header order.",
                i + 2, bmc_raw,
            )

        device = Device(
            row_index=i + 2,
            device_group=_get("device_group"),
            device_name=_get("device_name"),
            bmc_ip=bmc_raw,
            enabled=_bool(enabled_raw) if enabled_raw else True,
            bmc_username=_get("bmc_username"),
            bmc_password=_get("bmc_password"),
            inband_ip=_get("inband_ip"),
            inband_username=_get("inband_username"),
            inband_password=_get("inband_password"),
            tags=tags_str,
        )
        devices.append(device)

    return devices


def _load_task_defs(tasks_json_path: str | Path | None = None) -> dict[str, dict]:
    """Load task execution definitions from tasks.json.

    Searches multiple paths:
      1. Explicit tasks_json_path parameter
      2. Source root: <project>/tasks.json
      3. Packaged exe: <project>/_internal/tasks.json
      4. CWD-relative: ./tasks.json
    """
    if tasks_json_path is not None:
        candidates = [Path(tasks_json_path)]
    else:
        project_root = Path(__file__).resolve().parent.parent.parent
        cwd = Path.cwd()

        # Derive release root from sys.executable for PyInstaller onefile exe.
        # In onefile mode, sys.executable is the exe path (e.g. runtime/bmc-engine.exe),
        # and the release root is the parent of runtime/.
        _exe_dir = None
        try:
            import sys as _sys
            _exe_path = Path(_sys.executable).resolve()
            # If exe is in a 'runtime' dir, the release root is its parent
            if _exe_path.parent.name in ("runtime",):
                _exe_dir = _exe_path.parent.parent
        except Exception:
            pass

        candidates = []
        # Source layout: <project>/tasks.json
        candidates.append(project_root / "tasks.json")
        # Packaged exe: tasks.json at release root (alongside runtime/)
        candidates.append(cwd / "tasks.json")
        candidates.append(cwd / "app" / "tasks.json")
        # Release root derived from exe location
        if _exe_dir is not None:
            candidates.append(_exe_dir / "tasks.json")
            candidates.append(_exe_dir / "app" / "tasks.json")
        # Legacy: inside _internal/
        candidates.append(project_root / "_internal" / "tasks.json")
        candidates.append(cwd / "_internal" / "tasks.json")
        # Upward search: parent levels of project_root
        for ancestor in [project_root.parent, project_root.parent.parent]:
            candidates.append(ancestor / "tasks.json")
            candidates.append(ancestor / "app" / "tasks.json")

    found_path = None
    for p in candidates:
        if os.path.exists(str(p)):
            found_path = p
            break

    if found_path is None:
        logger.warning(
            "tasks.json not found in any of: %s, using Excel fallback — "
            "SSH tasks may have empty commands",
            [str(p) for p in candidates],
        )
        return {}

    with open(str(found_path), "r", encoding="utf-8") as f:
        data = json.load(f)

    tasks_def = data.get("tasks", {})
    if not isinstance(tasks_def, dict):
        logger.warning("tasks.json 'tasks' must be an object, got %s", type(tasks_def).__name__)
        return {}

    try:
        from ..rulepacks import merge_rule_packs_into_task_defs
        tasks_def = merge_rule_packs_into_task_defs(
            tasks_def,
            workspace_root=str(found_path.parent),
        )
    except Exception as e:
        logger.warning("RulePack merge skipped: %s", e)

    indexed: dict[str, dict] = {}
    for key, raw_def in tasks_def.items():
        if not isinstance(raw_def, dict):
            continue
        tdef = dict(raw_def)
        tdef.setdefault("_config_key", key)
        task_id = _str(tdef.get("task_id"))
        task_name = _str(tdef.get("task_name"))

        indexed[key] = tdef
        if task_id:
            indexed.setdefault(task_id, tdef)
        if task_name:
            indexed.setdefault(task_name, tdef)

    logger.info(
        "Loaded %d task definitions (%d lookup aliases) from %s",
        len(tasks_def), len(indexed), found_path,
    )
    return indexed


def _lookup_task_def(task_defs: dict[str, dict], task_id: str, task_name: str) -> tuple[dict, str]:
    """Resolve a task definition by stable task_id first, then legacy task_name."""
    task_id = _str(task_id)
    task_name = _str(task_name)
    if task_id:
        direct = task_defs.get(task_id)
        if isinstance(direct, dict):
            return direct, "task_id"
        for raw in task_defs.values():
            if isinstance(raw, dict) and _str(raw.get("task_id")) == task_id:
                return raw, "task_id"
    if task_name:
        direct = task_defs.get(task_name)
        if isinstance(direct, dict):
            return direct, "task_name"
        for raw in task_defs.values():
            if isinstance(raw, dict) and _str(raw.get("task_name")) == task_name:
                return raw, "task_name"
    return {}, ""


def load_tasks(
    filepath: str | Path,
    sheet_name: str = "任务列表",
    tasks_json_path: str | Path | None = None,
) -> list[Task]:
    """Load tasks from Excel, merging with tasks.json definitions by task_id."""
    task_defs = _load_task_defs(tasks_json_path)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    ws = wb[sheet_name]
    # Read header to find optional columns
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    task_headers = [_str(v) for v in header_row] if header_row else []
    task_col_map = _build_task_header_map(task_headers)
    full_screenshot_col = task_col_map.get("full_screenshot", -1)
    screenshot_mode_col = task_col_map.get("screenshot_mode", -1)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    tasks: list[Task] = []
    for i, row in enumerate(rows):
        if all(v is None for v in row):
            continue

        vals = [_str(v) for v in row]
        has_task_header_map = bool(task_col_map)

        def _task_val(field: str, fallback_idx: int = -1) -> str:
            idx = task_col_map.get(field, fallback_idx)
            if idx >= 0 and idx < len(vals):
                return vals[idx]
            return ""

        task_id = _task_val("task_id")
        task_name = _task_val("task_name", 1)

        # Resolve execution details: JSON takes priority, Excel is fallback
        tdef, matched_by = _lookup_task_def(task_defs, task_id, task_name)
        if matched_by == "task_name" and task_id:
            logger.warning(
                "Task row %d task_id=%s did not directly match tasks.json; "
                "matched by legacy task_name=%s instead",
                i + 2, task_id, task_name,
            )
        elif matched_by == "task_name":
            logger.warning(
                "Task row %d has no task_id; matched tasks.json by legacy task_name=%s",
                i + 2, task_name,
            )
        resolved_task_id = task_id or _str(tdef.get("task_id")) or _str(tdef.get("_config_key"))
        if not task_name and tdef:
            task_name = _str(tdef.get("task_name")) or task_name

        task_type = tdef.get("task_type") or _task_val("task_type", 2)
        execution_mode = tdef.get("execution_mode") or ""
        command_or_url = tdef.get("command_or_url") or ""
        actions_json = tdef.get("actions_json") or ""
        timeout_seconds = tdef.get("timeout_seconds", 60)
        retry_count = tdef.get("retry_count", 0)
        rules_json = json.dumps(tdef.get("rules", [])) if tdef.get("rules") else ""
        per_group_commands = tdef.get("per_group_commands") or {}
        per_group_commands_json = json.dumps(per_group_commands, ensure_ascii=False) if per_group_commands else ""

        # Excel columns depend on format:
        # Simplified (7 cols):  seq | name | type | group | outdir | imgname | enabled
        # Simplified with task_id (8+ cols): id | seq | name | type | group | outdir | imgname | enabled ...
        # Extended (9 cols):    seq | name | type | group | outdir | imgname | enabled | full_ss | ss_mode
        # Full legacy (14 cols): seq | name | type | group | mode | cmd | actions | outdir | imgname | timeout | retry | enabled | rules
        has_legacy_exec_fields = any(
            field in task_col_map
            for field in ("execution_mode", "command_or_url", "actions_json", "timeout_seconds", "retry_count", "rules_json")
        )
        is_simplified_or_extended = (
            not has_legacy_exec_fields if has_task_header_map else len(vals) <= 9
        )

        # Column mapping diagnostics
        if col_count := len(vals):
            logger.info(
                "任务行 %d [%s]: %d列 → "
                "task_id=%s seq=%s name=%s type=%s group=%s "
                "outdir=%s imgname=%s enabled=%s"
                "%s%s",
                i + 2, task_name, col_count,
                task_id,
                _task_val("sequence", 0),
                task_name,
                task_type,
                _task_val("match_group", 3),
                _task_val("output_dir_template", 4),
                _task_val("image_name_template", 5),
                _task_val("enabled", 6),
                f" full_ss[7]={vals[7]}" if col_count > 7 else "",
                f" ss_mode[8]={vals[8]}" if col_count > 8 else "",
            )

        if is_simplified_or_extended:
            match_group = _task_val("match_group", 3)
            output_dir_template = _task_val("output_dir_template", 4) or "{device_group}/{device_name}/{task_name}"
            image_name_template = _task_val("image_name_template", 5) or "{device_name}_{task_name}_{timestamp}"
            enabled_raw = _task_val("enabled", 6)
            enabled = _bool(enabled_raw) if enabled_raw else True
            # Task definition overrides: actions_json, enabled (from JSON)
            actions_json = tdef.get("actions_json") or ""
            # Only allow JSON to force-DISABLE (security gate).
            # Excel is the user-facing toggle; JSON must not force-enable.
            if tdef.get("enabled") is False:
                enabled = False
            if not tdef:
                # P0-3: no tasks.json definition → simplified format has no command column
                # execution_mode from vals[1] is WRONG (it's the task name, not exec mode)
                # For SSH/SSH_CMD tasks without commands, this produces empty-command tasks
                # that SSH executor would connect but execute nothing → EXEC_SUCCESS (bug).
                execution_mode = task_type  # Use task_type as fallback (SSH→SSH_CMD, BMC→BMC_URL)
                if task_type in ("SSH", "TELNET") and not command_or_url:
                    enabled = False
                    logger.error(
                        "P0-3: tasks.json missing for SSH task '%s' — no command available, "
                        "disabling task to prevent false EXEC_SUCCESS", task_name,
                    )
                elif task_type in ("BMC",) and not command_or_url:
                    logger.warning(
                        "No tasks.json definition for BMC task '%s', "
                        "and no command_or_url — task may fail", task_name,
                    )
                else:
                    logger.warning(
                        "No tasks.json definition for '%s', task may not execute correctly", task_name,
                    )
        else:
            # Legacy full-column format — Excel provides everything as before
            match_group = _task_val("match_group", 3)
            if not execution_mode:
                execution_mode = _task_val("execution_mode", 4)
            if not command_or_url:
                command_or_url = _task_val("command_or_url", 5)
            if not actions_json:
                actions_json = _task_val("actions_json", 6)
            if not tdef.get("timeout_seconds"):
                timeout_seconds = _int(_task_val("timeout_seconds", 9), 60)
            if not tdef.get("retry_count"):
                retry_count = _int(_task_val("retry_count", 10), 0)
            output_dir_template = _task_val("output_dir_template", 7) or "{device_group}/{device_name}/{task_name}"
            image_name_template = _task_val("image_name_template", 8) or "{device_name}_{task_name}_{timestamp}"
            enabled_raw = _task_val("enabled", 11)
            enabled = _bool(enabled_raw) if enabled_raw else True

        # CUSTOM_SCRIPT tasks are always disabled unless explicitly enabled in JSON
        if execution_mode == "CUSTOM_SCRIPT" and not tdef.get("enabled", False):
            enabled = False

        raw_seq = _task_val("sequence", 0) or str(i + 1)
        # Parse optional header-based columns
        raw_full_ss = _str(row[full_screenshot_col]) if full_screenshot_col >= 0 and len(row) > full_screenshot_col else ""
        full_ss = raw_full_ss.strip().lower() in ("是", "true", "1", "yes", "y")
        raw_ss_mode = _str(row[screenshot_mode_col]) if screenshot_mode_col >= 0 and len(row) > screenshot_mode_col else ""

        task = Task(
            row_index=i + 2,
            sequence=_int(raw_seq, i + 1),
            sequence_str=raw_seq,
            task_name=task_name,
            task_type=task_type,
            match_group=match_group,
            execution_mode=execution_mode,
            command_or_url=command_or_url,
            actions_json=actions_json,
            output_dir_template=output_dir_template,
            image_name_template=image_name_template,
            timeout_seconds=timeout_seconds,
            retry_count=retry_count,
            enabled=enabled,
            full_screenshot=full_ss,
            screenshot_mode=raw_ss_mode.strip() or "auto",
            rules_json=rules_json,
            task_id=resolved_task_id,
        )
        # Attach tasks.json definition for runtime lookup (checkpoints, ready conditions, etc.)
        if tdef:
            object.__setattr__(task, '_task_def', tdef)
        # Attach per_group_commands
        if per_group_commands_json:
            object.__setattr__(task, '_per_group_commands', per_group_commands)
        # Attach per_group_no_split
        per_group_no_split = tdef.get("per_group_no_split") or {}
        if per_group_no_split:
            object.__setattr__(task, '_per_group_no_split', per_group_no_split)
        # Attach per_group_timeout_seconds
        per_group_timeout_seconds = tdef.get("per_group_timeout_seconds") or tdef.get("per_group_timeout") or {}
        if per_group_timeout_seconds:
            object.__setattr__(task, '_per_group_timeout_seconds', per_group_timeout_seconds)
        # Attach global no_split
        if tdef.get("no_split"):
            object.__setattr__(task, '_no_split', True)
        tasks.append(task)

    return tasks


def load_all(
    filepath: str | Path,
    device_sheet: str = "设备信息",
    task_sheet: str = "任务列表",
    tasks_json_path: str | Path | None = None,
) -> tuple[list[Device], list[Task]]:
    return (
        load_devices(filepath, device_sheet),
        load_tasks(filepath, task_sheet, tasks_json_path),
    )
