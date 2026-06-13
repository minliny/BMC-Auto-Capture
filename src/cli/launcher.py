"""
BMC Auto Capture - 统一启动入口 (Launcher)
"""
from __future__ import annotations

import os
import sys
import json
import time
import socket
import argparse
import subprocess
import re
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    from .args import resolve_execution_cli
except ImportError:  # pragma: no cover - direct script execution fallback
    from src.cli.args import resolve_execution_cli

# 尝试导入 openpyxl
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 中文适配
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# 日志前缀常量
LOG_PREFIX_OK = "[成功]"
LOG_PREFIX_FAIL = "[失败]"
LOG_PREFIX_WARN = "[警告]"
LOG_PREFIX_INFO = "[信息]"
LOG_PREFIX_SKIP = "[跳过]"
LOG_PREFIX_PROC = "[处理]"


def log(msg: str, level: str = "INFO"):
    """带前缀的日志输出"""
    prefixes = {
        "INFO": LOG_PREFIX_INFO,
        "SUCCESS": LOG_PREFIX_OK,
        "FAIL": LOG_PREFIX_FAIL,
        "WARN": LOG_PREFIX_WARN,
        "SKIP": LOG_PREFIX_SKIP,
        "PROC": LOG_PREFIX_PROC,
    }
    prefix = prefixes.get(level, "[信息]")
    print(f"{prefix} {msg}")


def log_step(step: str, msg: str = ""):
    """步骤日志"""
    print(f"\n{'=' * 60}")
    print(f"  {step}" + (f" - {msg}" if msg else ""))
    print(f"{'=' * 60}")


def confirm(prompt: str, default: str = "N") -> str:
    """确认提示"""
    options = "Y/N/P"
    if default == "Y":
        options = "Y/N/P"
    elif default == "P":
        options = "Y/N/P"

    while True:
        response = input(f"{prompt} ({options}): ").strip().upper()
        if not response:
            return default
        if response in ("Y", "N", "P"):
            return response
        print(f"请输入 Y (是)、N (否) 或 P (预检查)")


def is_release_package() -> bool:
    """判断是否为 release 包"""
    base_dir = Path(__file__).resolve().parent.parent.parent
    exe_path = base_dir / "bmc-auto-capture"
    return exe_path.exists() or (base_dir / "bmc-auto-capture.exe").exists()


def find_python() -> Optional[str]:
    """查找 Python 解释器"""
    base_dir = Path(__file__).resolve().parent.parent.parent

    # release 包优先使用内置 python
    if is_release_package():
        python_path = base_dir / "_internal" / "Python3.framework" / "Versions" / "3.9" / "Resources" / "Python.app" / "Contents" / "MacOS" / "Python"
        if python_path.exists():
            return str(python_path)

    # 源码包查找 venv
    venv_paths = [
        base_dir / ".venv" / "bin" / "python",
        base_dir / ".venv" / "Scripts" / "python.exe",
        base_dir / "venv" / "bin" / "python",
        base_dir / "venv" / "Scripts" / "python.exe",
        base_dir / "offline_bmc_deps" / "python311" / "python.exe",
        base_dir / "python" / "python.exe",
    ]

    for path in venv_paths:
        if path.exists():
            return str(path)

    # 系统 python
    return sys.executable


def check_prerequisites() -> bool:
    """检查前置条件"""
    base_dir = Path(__file__).resolve().parent.parent.parent

    log_step("一、环境检测")

    # 检查目录
    log(f"当前目录: {base_dir}")

    # 检查 tasks.json
    tasks_path = base_dir / "tasks.json"
    if tasks_path.exists():
        log("tasks.json 存在", "SUCCESS")
    else:
        log("tasks.json 不存在", "FAIL")
        return False

    # 检查 Playwright 浏览器
    # P0: unified path — runtime/playwright_browsers is the canonical location
    # for both source and packaged exe environments.
    playwright_paths = [
        base_dir / "runtime" / "playwright_browsers",
        base_dir / "_internal" / "playwright",
    ]
    playwright_found = any(p.exists() for p in playwright_paths)
    if playwright_found:
        log("Playwright 浏览器资源存在", "SUCCESS")
    elif is_release_package():
        log("Playwright 浏览器资源未找到 (checked runtime/playwright_browsers, _internal/playwright)", "WARN")
    else:
        # 源码包检查系统 playwright
        python_exe = sys.executable
        result = subprocess.run(
            [python_exe, "-c", "from playwright.sync_api import sync_playwright; print('OK')"],
            capture_output=True,
            text=True,
            cwd=base_dir
        )
        if result.returncode == 0:
            log("Playwright 可用", "SUCCESS")
        else:
            log("Playwright 未安装", "WARN")

    # 检查输出目录
    output_dir = base_dir / "runtime" / "output"
    try:
        output_dir.mkdir(parents=True, exist_ok=True)
        # 测试写入
        test_file = output_dir / ".write_test"
        test_file.touch()
        test_file.unlink()
        log("输出目录可写", "SUCCESS")
    except Exception as e:
        log(f"输出目录不可写: {e}", "FAIL")
        return False

    return True


def check_high_risk_tasks(tasks_json_path: str, excel_path: Optional[str] = None) -> list[str]:
    """检查高风险任务"""
    high_risk_keywords = [
        "拔插硬盘后清除Foreign配置",
        "foreignConfig",
        "清除Foreign配置",
    ]

    found_risks = []

    # 检查 tasks.json
    try:
        with open(tasks_json_path, 'r', encoding='utf-8') as f:
            tasks_data = json.load(f)

        for task_name, task_def in tasks_data.get("tasks", {}).items():
            task_str = json.dumps(task_def, ensure_ascii=False).lower()
            for keyword in high_risk_keywords:
                if keyword.lower() in task_str:
                    found_risks.append(f"tasks.json: {task_name} ({keyword})")
    except Exception as e:
        log(f"检查 tasks.json 出错: {e}", "WARN")

    # 检查 Excel
    if excel_path and Path(excel_path).exists() and HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows(values_only=True):
                    row_str = str(row)
                    for keyword in high_risk_keywords:
                        if keyword in row_str:
                            found_risks.append(f"Excel: {sheet_name} ({keyword})")
            wb.close()
        except Exception as e:
            log(f"检查 Excel 出错: {e}", "WARN")

    return found_risks


def validate_excel(excel_path: str) -> dict:
    """验证 Excel 文件并返回摘要"""
    if not Path(excel_path).exists():
        raise FileNotFoundError(f"文件不存在: {excel_path}")

    if not excel_path.lower().endswith('.xlsx'):
        raise ValueError("文件扩展名必须是 .xlsx")

    if not HAS_OPENPYXL:
        raise ImportError("缺少 openpyxl 库，请安装: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # 获取任务列表
    task_sheet = wb.get_sheet_by_name("任务列表") if "任务列表" in wb.sheetnames else None
    device_sheet = wb.get_sheet_by_name("设备列表") if "设备列表" in wb.sheetnames else None

    task_count = 0
    enabled_task_count = 0
    device_count = 0
    enabled_device_count = 0

    if task_sheet:
        for row in task_sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            task_count += 1
            # 检查是否启用 (假设第7列或列名为"是否启用")
            enabled = True
            if len(row) >= 7 and row[6] in ("否", "禁用", "no", "false", "0"):
                enabled = False
            if enabled:
                enabled_task_count += 1

    if device_sheet:
        for row in device_sheet.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            device_count += 1
            # 检查是否启用
            enabled = True
            if len(row) >= 8 and row[7] in ("否", "禁用", "no", "false", "0"):
                enabled = False
            if enabled:
                enabled_device_count += 1

    wb.close()

    return {
        "task_count": task_count,
        "enabled_task_count": enabled_task_count,
        "device_count": device_count,
        "enabled_device_count": enabled_device_count,
    }


def check_network_connectivity(
    excel_path: str,
    timeout: int = 3,
    strict: bool = False
) -> dict:
    """检查网络连通性 — delegates to unified preflight implementation."""
    log_step("六、网络连通性检测")

    if not Path(excel_path).exists() or not HAS_OPENPYXL:
        log("无法检查网络连通性 (Excel 或 openpyxl 不可用)", "WARN")
        return {}

    try:
        from src.models.device import Device
        from src.connectivity.preflight import check_all as preflight_check_all

        devices = []
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        device_sheet = wb.get_sheet_by_name("设备列表") if "设备列表" in wb.sheetnames else None

        if device_sheet:
            for idx, row in enumerate(device_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if all(v is None for v in row):
                    continue
                if len(row) >= 8 and row[7] in ("否", "禁用", "no", "false", "0"):
                    continue
                device_name = row[1] if len(row) > 1 else ""
                device_group = row[0] if len(row) > 0 else ""
                bmc_ip = row[2] if len(row) > 2 else ""
                inband_ip = row[4] if len(row) > 4 else ""
                devices.append(Device(
                    row_index=idx,
                    device_name=device_name or "-",
                    device_group=device_group or "",
                    bmc_ip=str(bmc_ip or ""),
                    inband_ip=str(inband_ip or ""),
                    enabled=True,
                    bmc_username="",
                    bmc_password="",
                ))
        wb.close()

        if not devices:
            log("未找到启用的设备", "WARN")
            return {}

        report = preflight_check_all(devices, timeout=float(timeout))

        # Display results table
        print(f"\n{'设备名称':<20} {'设备分组':<15} {'BMC IP':<15} {'BMC':<8} {'带内IP':<15} {'SSH':<8}")
        print("-" * 85)
        for r in report.results:
            bmc_ip = next((d.bmc_ip for d in devices if d.device_name == r.device_name), "-")
            inband_ip = next((d.inband_ip for d in devices if d.device_name == r.device_name), "-")
            bmc_status = "可达" if r.bmc_status == "OK" else "不可达" if r.bmc_status else "-"
            ssh_status = "可达" if r.ssh_status == "OK" else "不可达" if r.ssh_status else "-"
            print(f"{r.device_name:<20} {r.device_group or '-':<15} {bmc_ip:<15} {bmc_status:<8} {inband_ip:<15} {ssh_status:<8}")
        print("-" * 85)
        print(f"共检测 {report.probe_count} 个端点，影响 {report.impacted_task_count} 个任务")

        return {"report": report}

    except Exception as e:
        log(f"网络检测出错: {e}", "WARN")
        return {}


def generate_output_dir(base_dir: Path, custom_dir: Optional[str] = None) -> Path:
    """生成输出目录"""
    if custom_dir:
        output_dir = base_dir / custom_dir
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = base_dir / "runtime" / "output" / timestamp

    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def summarize_plan(
    excel_summary: dict,
    output_dir: Path,
    concurrency: int,
    high_risks: list[str]
) -> None:
    """打印执行计划摘要"""
    log_step("七、执行计划确认")

    print(f"""
启用设备数量: {excel_summary.get('enabled_device_count', '?')}
启用任务数量: {excel_summary.get('enabled_task_count', '?')}
预计输出目录: {output_dir}
并发数: {concurrency}

风险检查: {'发现高风险任务' if high_risks else '未发现'}
""")

    if high_risks:
        print("高风险任务列表:")
        for risk in high_risks:
            print(f"  - {risk}")
        print()


def run_execution(
    excel_path: str,
    output_dir: Path,
    concurrency: int,
    mode: str,
    precheck_only: bool = False
) -> int:
    """执行 BMC 任务"""
    log_step("九、任务执行")

    base_dir = Path(__file__).resolve().parent.parent.parent

    # 构建命令 - 使用 run.py 兼容的参数格式
    if is_release_package():
        # release 包
        exe_path = base_dir / "bmc-auto-capture"
        if exe_path.exists():
            cmd = [str(exe_path)]
        else:
            exe_path = base_dir / "_internal" / "Python3.framework" / "Versions" / "3.9" / "Resources" / "Python.app" / "Contents" / "MacOS" / "Python"
            cmd = [str(exe_path), str(base_dir / "run.py")]
    else:
        # 源码包
        python_path = find_python()
        cmd = [python_path, str(base_dir / "run.py")]

    cmd.extend([
        "--excel", excel_path,
        "--mode", mode,
    ])

    # Pass CLI overrides to run.py
    if output_dir is not None:
        cmd.extend(["--output", str(output_dir)])
    if concurrency and concurrency > 1:
        # --concurrency deprecated: map missing worker pools for compatibility
        if args.max_bmc_workers is None:
            cmd.extend(["--max-bmc-workers", str(concurrency)])
            log(f"[兼容] --concurrency {concurrency} → 映射为 --max-bmc-workers {concurrency}", "WARN")
        if args.max_ssh_workers is None:
            cmd.extend(["--max-ssh-workers", str(concurrency)])
            log(f"[兼容] --concurrency {concurrency} → 映射为 --max-ssh-workers {concurrency}", "WARN")
    if args.max_bmc_workers is not None:
        cmd.extend(["--max-bmc-workers", str(args.max_bmc_workers)])
    if args.max_ssh_workers is not None:
        cmd.extend(["--max-ssh-workers", str(args.max_ssh_workers)])
    if args.bmc_artifact_profile is not None:
        cmd.extend(["--bmc-artifact-profile", args.bmc_artifact_profile])

    if precheck_only:
        cmd.append("--preflight-only")

    log(f"执行命令: {' '.join(cmd)}", "INFO")

    try:
        result = subprocess.run(
            cmd,
            cwd=str(base_dir),
            text=True,
            encoding='utf-8',
            errors='replace',
        )
        # 打印子进程输出
        if result.stdout:
            print(result.stdout)
        if result.stderr:
            print(result.stderr, file=sys.stderr)
        return result.returncode
    except Exception as e:
        log(f"执行出错: {e}", "FAIL")
        return -1


def summarize_results(output_dir: Path) -> None:
    """汇总执行结果"""
    log_step("十、结果汇总")

    result_csv = output_dir / "result.csv"
    if not result_csv.exists():
        log("未找到 result.csv", "WARN")
        return

    try:
        import csv
        with open(result_csv, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        total = len(rows)
        success = sum(1 for r in rows if r.get("执行状态") == "EXEC_SUCCESS")
        failed = sum(1 for r in rows if r.get("执行状态") == "EXEC_FAILED")
        partial = sum(1 for r in rows if r.get("执行状态") == "EXEC_PARTIAL")
        skipped = sum(1 for r in rows if "SKIPPED" in r.get("执行状态", ""))

        png_count = sum(
            len(r.get("截图", "").split(";"))
            for r in rows if r.get("截图")
        )
        html_count = sum(1 for r in rows if r.get("HTML文件"))
        txt_count = sum(1 for r in rows if r.get("TXT文件"))

        print(f"""
result.csv 路径: {result_csv}
总任务数: {total}
成功数: {success}
失败数: {failed}
部分成功数: {partial}
跳过数: {skipped}
PNG 数量: {png_count}
HTML 数量: {html_count}
TXT 数量: {txt_count}
""")

        # artifact_status 分布
        artifact_stats = {}
        for r in rows:
            status = r.get("工件状态", "N/A")
            artifact_stats[status] = artifact_stats.get(status, 0) + 1

        print("工件状态分布:")
        for status, count in sorted(artifact_stats.items()):
            print(f"  {status}: {count}")

        # ready_status 分布
        ready_stats = {}
        for r in rows:
            status = r.get("就绪状态", "N/A")
            ready_stats[status] = ready_stats.get(status, 0) + 1

        print("\n就绪状态分布:")
        for status, count in sorted(ready_stats.items()):
            print(f"  {status}: {count}")

    except Exception as e:
        log(f"读取结果出错: {e}", "WARN")


def main():
    """主入口"""
    parser = argparse.ArgumentParser(description="BMC Auto Capture - 统一启动入口")
    parser.add_argument("--excel", "-e", help="Excel 文件路径")
    parser.add_argument("--output", "-o", help="输出目录")
    parser.add_argument("--concurrency", "-c", type=int, default=1, help="(已弃用) 并发数 — 建议使用 --max-bmc-workers / --max-ssh-workers")
    parser.add_argument("--mode", choices=["sequential", "full"], default=None, help="执行模式：sequential 或 full")
    parser.add_argument("--max-bmc-workers", type=int, default=None, help="BMC 最大并发数")
    parser.add_argument("--max-ssh-workers", type=int, default=None, help="SSH 最大并发数")
    parser.add_argument("--bmc-artifact-profile", choices=["full", "fast"], default=None, help="BMC 证据模式：full 完整，fast 仅 PNG/HTML")
    parser.add_argument("--strict", action="store_true", help="严格模式 (网络检测失败则中止)")
    parser.add_argument("--precheck-only", action="store_true", help="仅预检查模式")
    parser.add_argument("--yes", "-y", action="store_true", help="跳过确认直接执行")

    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parent.parent.parent

    # 一、环境检测
    log_step("零、中文编码适配")
    log(f"Python: {sys.version}", "INFO")
    log(f"工作目录: {base_dir}", "INFO")
    log(f"包类型: {'Release' if is_release_package() else '源码'}", "INFO")

    if not check_prerequisites():
        log("前置检查失败", "FAIL")
        return 1

    # 二、Excel 配置
    log_step("三、Excel 路径配置")

    excel_path = args.excel
    if not excel_path:
        log("请输入 Excel 文件路径 (可直接拖拽文件到窗口)", "INFO")
        while True:
            excel_path = input("Excel 路径: ").strip().strip('"')
            if not excel_path:
                log("路径不能为空", "FAIL")
                continue
            if Path(excel_path).exists():
                break
            log(f"文件不存在: {excel_path}", "FAIL")

    try:
        excel_summary = validate_excel(excel_path)
        log(f"Excel 验证成功", "SUCCESS")
        log(f"设备总数: {excel_summary['device_count']}, 启用: {excel_summary['enabled_device_count']}", "INFO")
        log(f"任务总数: {excel_summary['task_count']}, 启用: {excel_summary['enabled_task_count']}", "INFO")
    except Exception as e:
        log(f"Excel 验证失败: {e}", "FAIL")
        return 1

    # 四、输出目录配置
    log_step("四、输出目录配置")
    output_dir = generate_output_dir(base_dir, args.output)
    log(f"输出目录: {output_dir}", "SUCCESS")

    # 五、风险检查
    log_step("五、风险任务检查")
    risks = check_high_risk_tasks(str(base_dir / "tasks.json"), excel_path)
    if risks:
        log("检测到高风险状态变更任务，已停止执行", "FAIL")
        log("请从任务配置中删除以下项后重试:", "FAIL")
        for risk in risks:
            print(f"  - {risk}")
        return 1
    else:
        log("未发现高风险任务", "SUCCESS")

    # 六、网络连通性检测
    check_network_connectivity(excel_path, timeout=3, strict=args.strict)

    # 并发配置
    log_step("八、并发配置")
    concurrency = args.concurrency
    if concurrency < 1:
        concurrency = 1
    args.concurrency = concurrency
    mode, resolved_bmc_workers, resolved_ssh_workers, concurrency = resolve_execution_cli(args, sys.argv[1:])
    log(f"并发数 (--concurrency, 已弃用): {concurrency}", "WARN")
    log(f"执行模式: {mode}", "INFO")
    if resolved_bmc_workers is not None:
        log(f"BMC 最大并发用户: {resolved_bmc_workers}", "INFO")
    if resolved_ssh_workers is not None:
        log(f"SSH 最大并发用户: {resolved_ssh_workers}", "INFO")
    if resolved_bmc_workers is None and resolved_ssh_workers is None:
        log(f"未指定 --max-bmc-workers / --max-ssh-workers，将使用 YAML config 默认值", "INFO")
    log("提示: 大规模运行建议 --max-bmc-workers 4 --max-ssh-workers 20", "INFO")

    # 七、执行计划确认
    summarize_plan(excel_summary, output_dir, concurrency, risks)

    if args.yes:
        response = "Y"
    else:
        response = confirm("是否开始执行? (Y=是, N=否, P=仅预检查)", "N")

    if response == "N":
        log("已取消执行", "WARN")
        return 0

    if response == "P":
        log("预检查模式，不执行任务", "SKIP")
        return 0

    # 九、任务执行
    return_code = run_execution(excel_path, output_dir, concurrency, mode, args.precheck_only)

    # 十、结果汇总
    if return_code == 0:
        summarize_results(output_dir)

    if return_code != 0:
        log(f"执行完成，退出码: {return_code}", "FAIL")
        log(f"请查看 result.csv: {output_dir / 'result.csv'}", "INFO")
    else:
        log("执行完成", "SUCCESS")

    return return_code


if __name__ == "__main__":
    sys.exit(main())
